'''
Author: spc
Email: hilqiqi0@foxmail.com
Date: 2024-04-11 20:12:26
LastEditors: spc
LastEditTime: 2024-04-11 20:19:43
FilePath: /stock/data_processing/show_limit_excel.py
Description: 

Copyright (c) 2024 by spc, All Rights Reserved. 
'''
from sqlalchemy import create_engine

engine_ts = create_engine('mysql://root@127.0.0.1:3306/tushare_feature?charset=utf8&use_unicode=1')

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill

# 创建 Workbook 对象
wb = Workbook()
# 选择默认的活动工作表
ws = wb.active

start_date = '20240401'
end_date = '20240410'
sql = f"""SELECT * 
          FROM tushare_stock.trade_cal 
          WHERE cal_date >= '{start_date}' and cal_date <= '{end_date}'
          order by cal_date asc 
          LIMIT 1000"""
df_cal = pd.read_sql_query(sql, engine_ts)

# print(df_cal)



sql = """SELECT * 
         FROM tushare_feature.limit_list_d 
         WHERE trade_date >= '{start_date}' and trade_date <= '{end_date}' 
         """
df_limit = pd.read_sql_query(sql, engine_ts)


# print(df)

trade_dates_sorted = sorted(set(df_limit['trade_date'].values), reverse=False)

# print(trade_dates_sorted)

# print(df[df['trade_date'] == '20240411'])

col_idx_map = {}
row_idx_map = {}
row_idx_num = 2
for col_idx, trade_date in enumerate(trade_dates_sorted, start=2):
    print(col_idx, trade_date)
    # 第一行：日期
    cell = ws.cell(row=1, column=col_idx, value=trade_date)

    col_idx_map[trade_date] = col_idx
    df_day = df_limit[df_limit['trade_date'] == trade_date]

    # 对涨停统计
    # df_up = df_day[df_day['limit'] == 'U']
    df_up = df_day
    value_counts_sorted = df_up['industry'].value_counts().sort_values(ascending=False)

    for industry, count in value_counts_sorted.items():
        print(industry, count)
        df_up_ts = df_up[df_up['industry'] == industry]
        for index, row in df_up_ts.iterrows():
            ts_name = row["name"]
            if not row_idx_map.get(ts_name):
                row_idx_map[ts_name] = row_idx_num
                # 第一列：行业
                cell = ws.cell(row=row_idx_num, column=1, value=industry)

                row_idx_num += 1

            limit = row["limit"]  # D跌停U涨停Z炸板
            limit_times = row["limit_times"]  # 连板数
            if col_idx_map[trade_date] == 2:
                # 第一列
                cell = ws.cell(row=row_idx_map.get(ts_name), column=col_idx_map[trade_date], value=ts_name)
            elif limit in ('D', 'Z'):
                # 跌停、炸板
                cell = ws.cell(row=row_idx_map.get(ts_name), column=col_idx_map[trade_date], value=ts_name)
            elif limit_times == 1.0:
                cell = ws.cell(row=row_idx_map.get(ts_name), column=col_idx_map[trade_date], value=ts_name)
            else:
                cell = ws.cell(row=row_idx_map.get(ts_name), column=col_idx_map[trade_date], value=limit_times)

            # 涨跌炸板
            if row["limit"] == 'U':
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif row["limit"] == 'Z':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            # print(industry, ts_name, row["limit_times"])
        # break

    # break

# print(col_idx_map)

wb.save('ts_output.xlsx')
#
#
# from openpyxl import Workbook
# from openpyxl.styles import Font, Color, PatternFill
#
# # 创建 Workbook 对象
# wb = Workbook()
# # 选择默认的活动工作表
# ws = wb.active
#
# # 数据列表
# data = [
#     ['Name', 'Age1', 'Country',"wwww","as"],
#     ['John', 25, 'USA'],
#     ['Emily', 30, 'Canada'],
#     ['Michael', 35, 'Australia'],
# ]
#
# # 写入数据和设置单元格样式
# for row_idx, row_data in enumerate(data, start=1):
#     for col_idx, cell_value in enumerate(row_data, start=1):
#         cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
#         cell.font = Font(color="FF0000", bold=True)
#         cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#
# # 保存 Excel 文件
# wb.save('output.xlsx')
